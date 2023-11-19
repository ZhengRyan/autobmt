#!/usr/bin/env python
# ! -*- coding: utf-8 -*-

'''
@File: plot.py
@Author: RyanZheng
@Email: ryan.zhengrp@gmail.com
@Created Time on: 2020-10-18
'''
import os
import re

import matplotlib.pyplot as plt
import numpy as np
import openpyxl
import pandas as pd
import six
from matplotlib import gridspec
from openpyxl.drawing.image import Image
from sklearn.metrics import confusion_matrix
from sklearn.metrics import roc_curve, roc_auc_score, precision_recall_curve

from .statistics import calc_var_summary
from .utils import save_json


def ellipsis_fun(x, decimal=2, ellipsis=16):
    if re.search(r'\[(.*?) ~ (.*?)\)', str(x)):
        le = x.split('~')[0]
        ri = x.split('~')[1]
        left = re.match(r'\d+.\[(.*?) ', le)
        right = re.match(r' (.*?)\)', ri)
        if left:
            tmp = round(float(left.groups()[0]), decimal)
            l = f"{le.split('[')[0]}[{tmp} "
            le = l
        if right:
            tmp = round(float(right.groups()[0]), decimal)
            r = f" {tmp})"
            ri = r

        return f"{le} ~ {ri}"
    else:
        return str(x)[:ellipsis] + '..'


def plot_var_bin_summary(frame, cols, target='target', by='type', file_path=None, sheet_name='plot_var_bin_summary',
                         need_bin=False, decimal=2, ellipsis=16, **kwargs):
    """
    画变量分箱图
    Args:
        frame (DataFrame):
        cols (str|list): 需要画图的变量
        target (str): 目标变量
        by (str): 分开统计的列名
        file_path (str): 保存路径
        sheet_name (str):保存excel的sheet名称
        need_bin (bool):是否要进行分箱

    Returns:

    """

    if isinstance(cols, str):
        cols = [cols]

    if isinstance(frame, pd.DataFrame):
        summary_df_dict = {}
        if by in frame:
            for name, df in frame.groupby(by):
                summary_df_dict[name] = calc_var_summary(df, include_cols=cols, target=target, need_bin=need_bin,
                                                         **kwargs)
        else:
            summary_df_dict['all'] = calc_var_summary(frame, include_cols=cols, target=target, need_bin=need_bin,
                                                      **kwargs)
    else:
        summary_df_dict = frame
    summary_num = len(summary_df_dict)

    save_jpg_path = None
    if file_path is not None:
        if '.xlsx' in file_path:
            fp_sp = file_path.split('.xlsx')
            if '' == fp_sp[1]:
                save_jpg_path = fp_sp[0] + '_var_jpg'
                os.makedirs(save_jpg_path, exist_ok=True)

            else:
                save_jpg_path = os.path.join(file_path, 'var_jpg')
                os.makedirs(save_jpg_path, exist_ok=True)
                file_path = os.path.join(file_path, 'plot_var_bin_summary.xlsx')
        else:
            save_jpg_path = os.path.join(file_path, 'var_jpg')
            os.makedirs(save_jpg_path, exist_ok=True)
            file_path = os.path.join(file_path, 'plot_var_bin_summary.xlsx')

    for col in cols:
        # 做图
        # plt.figure(figsize=(15, 4))
        # gs = gridspec.GridSpec(1, 2)
        # gs = gridspec.GridSpec(1, 3)

        gs_num = len(summary_df_dict)

        if gs_num == 1:
            plt.figure(figsize=(gs_num * 5, 4), dpi=400)
        else:
            plt.figure(figsize=(gs_num * 5, 4))

        gs = gridspec.GridSpec(1, gs_num)

        for i, k in enumerate(summary_df_dict):

            df = summary_df_dict[k]

            if i == 0:
                ax1 = plt.subplot(gs[0, 0])
            else:
                ax1 = plt.subplot(gs[0, i], sharey=ax1)

            df = df[df['var_name'] == col]
            df['range'] = df['range'].map(lambda x: ellipsis_fun(x, decimal, ellipsis))

            x_point = np.arange(len(df))
            y_point = df['total_pct']

            ax1.bar(x_point, y_point, color='Orange', alpha=0.4, width=0.5, label='PctTotal')
            x_labels = list(df['range'])
            plt.xticks(np.arange(len(df)), x_labels, fontsize=10, rotation=45)

            for x, y in zip(x_point, y_point):
                ax1.text(x + 0.05, y + 0.01, str(round(y * 100, 2)) + '%', ha='center', va='bottom', fontsize=12)
            ax1.set_ylabel('total_pct', fontsize=12)

            ax1.set_ylim([0, max(y_point) + ((max(y_point) - min(y_point)) / len(y_point))])
            bottom, top = ax1.get_ylim()
            ax2 = ax1.twinx()
            ax2.plot(x_point, df['positive_rate'], '-ro', color='red')

            for x, y in zip(x_point, df['positive_rate']):
                ax2.text(x + 0.05, y, str(round(y * 100, 2)) + '%', ha='center', va='bottom', fontsize=12, color='r')
            ax2.set_ylabel('positive_rate', fontsize=12)
            # ax2.set_ylim([0, max(df['positive_rate']) + 0.01])
            ax2.set_ylim(
                [0, max(df['positive_rate']) + ((max(df['positive_rate']) - min(df['positive_rate'])) / len(df))])

            plt.title('{}:{}\nIV: {:.5f}'.format(k, col, df['IV'].iloc[0]), loc='right', fontsize='small')
            # 将数据详情表添加
            tmp_df = df[['range', 'woe', 'iv', 'total']]
            round_cols = ['woe', 'iv']
            tmp_df[round_cols] = tmp_df[round_cols].applymap(lambda v: round(v, 4) if pd.notnull(v) else '')
            mpl_table = plt.table(cellText=tmp_df.values, colLabels=tmp_df.columns,
                                  colWidths=[0.2, 0.1, 0.1, 0.1],
                                  loc='top')  # loc='top'将详情放到顶部

            mpl_table.auto_set_font_size(False)
            mpl_table.set_fontsize(5)

            header_color = 'darkorange'
            row_colors = ['bisque', 'w']
            header_columns = 0

            for k, cell in six.iteritems(mpl_table._cells):
                cell.set_edgecolor('w')
                if k[0] == 0 or k[1] < header_columns:
                    cell.set_text_props(weight='bold', color='w')
                    cell.set_facecolor(header_color)
                else:
                    cell.set_facecolor(row_colors[k[0] % len(row_colors)])

        plt.tight_layout()

        if save_jpg_path is not None:  # 判断是否需要保存
            plt.savefig(os.path.join(save_jpg_path, '{}.png'.format(col)), dpi=300, bbox_inches='tight')  # dpi控制清晰度
        # plt.show()

    # if save_jpg_path is not None:
    #     workbook = xlsxwriter.Workbook(file_path)
    #     worksheet = workbook.add_worksheet(sheet_name)
    #
    #     for i, jpg_name in enumerate(cols):
    #         worksheet.insert_image('A{}'.format(i * 29 + 3), os.path.join(save_jpg_path, '{}.jpg'.format(jpg_name),
    #                                {'x_scale': 1.5, 'y_scale': 1.5})
    #
    #     workbook.close()
    if save_jpg_path is not None:
        try:
            wb = openpyxl.load_workbook(file_path)
        except:
            wb = openpyxl.Workbook()
        sh = wb.create_sheet(sheet_name)
        for i, jpg_name in enumerate(cols):
            img = Image(os.path.join(save_jpg_path, '{}.png'.format(jpg_name)))
            newsize = (summary_num * 700, 600)
            img.width, img.height = newsize  # 设置图片的宽和高
            sh.add_image(img, 'A{}'.format(i * 35 + 3))
        wb.save(file_path)


def get_optimal_cutoff(fpr_recall, tpr_precision, threshold, is_f1=False):
    if is_f1:
        youdenJ_f1score = (2 * tpr_precision * fpr_recall) / (tpr_precision + fpr_recall)
    else:
        youdenJ_f1score = tpr_precision - fpr_recall
    point_index = np.argmax(youdenJ_f1score)
    optimal_threshold = threshold[point_index]
    point = [fpr_recall[point_index], tpr_precision[point_index]]
    return optimal_threshold, point


def plot_ks(y_true: pd.Series, y_pred: pd.Series, output_path='', title=''):
    fpr, tpr, thresholds = roc_curve(y_true, y_pred)
    ks = max(tpr - fpr)  ###计算ks的值

    plt.figure(figsize=(6, 6))
    x = np.arange(len(thresholds)) / len(thresholds)
    plt.plot(x, tpr, lw=1)
    plt.plot(x, fpr, lw=1)
    plt.plot(x, tpr - fpr, lw=1, linestyle='--', label='KS curve (KS = %0.4f)' % ks)

    optimal_th, optimal_point = get_optimal_cutoff(fpr_recall=fpr, tpr_precision=tpr, threshold=thresholds)
    optimal_th_index = np.where(thresholds == optimal_th)
    plt.plot(optimal_th_index[0][0] / len(thresholds), ks, marker='o', color='r')
    plt.text(optimal_th_index[0][0] / len(thresholds), ks, (float('%.4f' % optimal_point[0]),
                                                            float('%.4f' % optimal_point[1])),
             ha='right', va='top', fontsize=12)
    plt.text(optimal_th_index[0][0] / len(thresholds), ks, f'threshold:{optimal_th:.4f}', fontsize=12)

    plt.xlim([0.0, 1.0])
    plt.ylim([0.0, 1.0])
    plt.xlabel('Thresholds Index')
    plt.ylabel('TPR FPR KS')
    name = '{} KS Curve'.format(title)
    plt.title(name)
    plt.legend(loc="lower right", fontsize=12)
    plt.savefig(output_path + name, bbox_inches='tight')
    plt.show()
    return {'KS': ks, 'KS最大值-threshold': optimal_th}


def plot_roc(y_true: pd.Series, y_pred: pd.Series, output_path='', title=''):
    fpr, tpr, thresholds = roc_curve(y_true, y_pred)
    auc_value = roc_auc_score(y_true, y_pred)  ###计算auc的值

    lw = 2
    plt.figure(figsize=(6, 6))

    plt.plot(fpr, tpr, color='darkorange',
             lw=lw, label='ROC curve (AUC = %0.4f)' % auc_value)

    plt.plot([0, 1], [0, 1], color='navy', lw=lw, linestyle='--')

    optimal_th, optimal_point = get_optimal_cutoff(fpr_recall=fpr, tpr_precision=tpr, threshold=thresholds)
    plt.plot(optimal_point[0], optimal_point[1], marker='o', color='r')
    plt.text(optimal_point[0], optimal_point[1], (float('%.4f' % optimal_point[0]),
                                                  float('%.4f' % optimal_point[1])),
             ha='right', va='top', fontsize=12)
    plt.text(optimal_point[0], optimal_point[1], f'threshold:{optimal_th:.4f}', fontsize=12)

    plt.xlim([0.0, 1.0])
    plt.ylim([0.0, 1.0])
    plt.xlabel('False Positive Rate(FPR)')
    plt.ylabel('True Positive Rate(TPR)')
    name = '{} ROC Curve'.format(title)
    plt.title(name)
    plt.legend(loc="lower right", fontsize=12)
    plt.savefig(output_path + name, bbox_inches='tight')
    plt.show()
    return {'AUC': auc_value}


def plot_pr(y_true: pd.Series, y_pred: pd.Series, output_path='', title=''):
    precision, recall, thresholds = precision_recall_curve(y_true, y_pred)
    f1score = (2 * precision * recall) / (precision + recall)  ###计算F1score
    max_f1score = max(f1score)

    lw = 2
    plt.figure(figsize=(6, 6))

    plt.plot(recall, precision, color='darkorange',
             lw=lw, label='PR curve (F1score = %0.4f)' % max_f1score)

    plt.plot([0, 1], [0, 1], color='navy', lw=lw, linestyle='--')

    optimal_th, optimal_point = get_optimal_cutoff(fpr_recall=recall, tpr_precision=precision, threshold=thresholds,
                                                   is_f1=True)
    plt.plot(optimal_point[0], optimal_point[1], marker='o', color='r')
    plt.text(optimal_point[0], optimal_point[1], (float('%.4f' % optimal_point[0]),
                                                  float('%.4f' % optimal_point[1])),
             ha='right', va='top', fontsize=12)
    plt.text(optimal_point[0], optimal_point[1], f'threshold:{optimal_th:.4f}', fontsize=12)

    plt.xlim([0.0, 1.0])
    plt.ylim([0.0, 1.0])
    plt.xlabel('Recall')
    plt.ylabel('Precision')
    name = '{} PR Curve'.format(title)
    plt.title(name)
    plt.legend(loc="lower right", fontsize=12)
    plt.savefig(output_path + name, bbox_inches='tight')
    plt.show()
    return {'F1_Score最大值': max_f1score, 'F1_Score最大值-threshold': optimal_th, '模型拐点': optimal_th, '阀值': optimal_th,
            'Precision': optimal_point[1], 'Recall': optimal_point[0], 'F1_Score': max_f1score}


def plot_pr_f1(y_true: pd.Series, y_pred: pd.Series, output_path='', title=''):
    precision, recall, thresholds = precision_recall_curve(y_true, y_pred)
    thresholds = np.insert(thresholds, 0, 0, axis=None)
    f1score = (2 * precision * recall) / (precision + recall)  ###计算F1score

    x = np.arange(len(thresholds)) / len(thresholds)

    pr_f1_dict = {'Precision': precision, 'Recall': recall, 'F1_score': f1score}

    for i in pr_f1_dict:
        plt.figure(figsize=(6, 6))

        plt.plot(x, pr_f1_dict[i], lw=1)

        plt.xlim([0.0, 1.0])
        plt.ylim([0.0, 1.0])
        plt.xlabel('Thresholds Index')
        plt.ylabel('{}'.format(i))
        name = '{} {} Curve'.format(title, i)
        plt.title(name)
        plt.savefig(output_path + name, bbox_inches='tight')
        plt.show()

    return {'Thresholds': list(thresholds), '模型召回率': list(recall), '模型精准率': list(precision),
            '模型F1-score': list(f1score)}


def calc_celue_cm(df: pd.DataFrame, target='target', to_bin_col='p'):
    q_cut_list = np.arange(0, 1, 1 / 10) + 0.1
    confusion_matrix_df = pd.DataFrame()
    for i in q_cut_list:
        df['pred_label'] = np.where(df[to_bin_col] >= i, 1, 0)
        tmp_list = []
        tmp_list.append(i)

        tn, fp, fn, tp = confusion_matrix(np.array(df[target]), np.array(df['pred_label'])).ravel()

        tmp_list.extend([tp, fp, tn, fn])

        confusion_matrix_df = confusion_matrix_df.append(pd.DataFrame(tmp_list).T)

    # confusion_matrix_df.columns = ['阈值', 'TP', 'FP', 'TN', 'FN']
    confusion_matrix_df.columns = ['阈值', '实际正样本-预测为正样本', '实际负样本-预测为正样本', '实际负样本-预测为负样本', '实际正样本-预测为负样本']
    confusion_matrix_df.set_index('阈值', inplace=True)
    confusion_matrix_df['sum'] = confusion_matrix_df.apply(lambda x: x.sum(), axis=1)

    # return confusion_matrix_df
    return confusion_matrix_df.to_dict()


def calc_plot_metrics(df: pd.DataFrame, to_bin_col='p', target='target', curve_save_path=''):
    data = {k: v for k, v in df.groupby('type')}
    data.update({'all': df})

    for data_type, type_df in data.items():
        res_save_path = os.path.join(curve_save_path, data_type)
        os.makedirs(res_save_path, exist_ok=True)

        res_dict = {}
        res_dict.update(plot_roc(type_df[target], type_df[to_bin_col], res_save_path, data_type))
        res_dict.update(plot_ks(type_df[target], type_df[to_bin_col], res_save_path, data_type))
        res_dict.update(plot_pr(type_df[target], type_df[to_bin_col], res_save_path, data_type))
        res_dict.update(calc_celue_cm(type_df, target, to_bin_col))
        res_dict.update(plot_pr_f1(type_df[target], type_df[to_bin_col], res_save_path, data_type))

        ###相关指标保存json格式
        save_json(res_dict, os.path.join(res_save_path, '{}_res_json.json'.format(data_type)))


if __name__ == "__main__":
    ######读取数据
    data_path = 'TD47p25combine_td_to_report_data.csv'
    df = pd.read_csv(data_path)
    df = df[df['label'].notnull()]

    print(len(df))

    ######结果保存路径
    curve_save_path = '../examples/curve_result'
    calc_plot_metrics(df=df, to_bin_col='td', target='label', curve_save_path=curve_save_path)
